from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def fetch_data_from_excel(file_path, sheet_name=None):
    # Load the workbook
    workbook = load_workbook(filename=file_path, data_only=True)
    
    # Select the sheet by name or the active sheet by default
    if sheet_name:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active
    
    # Fetch data as a list of lists (rows and columns)
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    
    return data

def fetch_data_from_sheet(workbook, sheet_name):    
    # Select the sheet by name or the active sheet by default
    if sheet_name:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active
    
    # Fetch data as a list of lists (rows and columns)
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    
    return data

def excel_save(file_path):
    workbook.save(file_path)

def update_excel_data(workbook, sheet_name, row, column, new_value):
    sheet = workbook[sheet_name]
    # Update the specific cell
    sheet.cell(row=row+1, column=column+1, value=new_value)    



def change_row_color(workbook, sheet_name, row, color):
    sheet = workbook[sheet_name]
    
    # Create a fill pattern with the specified color
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    # Apply the color to all cells in the specified row
    for i in range(10):
      cell = sheet.cell(row=row + 1, column=i+1)
      cell.fill = fill
    print(f"row {row} color {color}")

def change_cell_color(workbook, sheet_name, row, column, color):
    sheet = workbook[sheet_name]
    
    # Create a fill pattern with the specified color
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    # Apply the color to the specified cell
    cell = sheet.cell(row=row + 1, column=column + 1)
    cell.fill = fill
    
    
