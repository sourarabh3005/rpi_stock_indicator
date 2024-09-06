# Import the functions from excel_utils.py
from excel_utils import fetch_data_from_excel, update_excel_data, change_cell_color, fetch_data_from_sheet, change_row_color
from gdrive import upload_file_to_gdrive, download_file_from_gdrive
from enum import Enum
from sheet import SystemFields,  StkWishList, StkPortfolio
import shutil
from openpyxl import load_workbook
from datetime import datetime
import yfinance as yf
from  task_def import TASK_SYSTEM_DEFAULT, TASK_SYSTEM_RUNNING
from task_def import TASK_SYSTEM_STK_BUY, TASK_SYSTEM_STK_SELL, TASK_SYSTEM_STK_CRT
from task_def import TASK_SYSTEM_STK_BUY_CLR, TASK_SYSTEM_STK_SELL_CLR, TASK_SYSTEM_STK_CRT_CLR

MAX_ALLOWED_ROWS = 100

def get_stock_price(ticker_name, tkr_type):
    print(f"type {tkr_type}")
    if tkr_type == "IND":
      ticker_name = ticker_name + '.NS'
    try:
        # Fetch the ticker data
        ticker = yf.Ticker(ticker_name)
        # Get the current stock price
        stock_price = ticker.history(period="1d")['Close'].iloc[-1]
        return stock_price
    except IndexError:
        return "Error: Invalid ticker name or no data available."
    except Exception as e:
        return f"An error occurred: {e}"

def copy_file(src, dest):
    """
    Copy a file from src to dest.

    :param src: Source file path
    :param dest: Destination file path
    """
    try:
        shutil.copy(src, dest)
        print(f"File copied successfully from {src} to {dest}.")
    except FileNotFoundError:
        print(f"Source file not found: {src}")
    except PermissionError:
        print(f"Permission denied. Cannot copy file to {dest}.")
    except Exception as e:
        print(f"An error occurred: {e}")
        
def get_current_date():
    """
    Returns the current date in the format YYYY-MM-DD.
    """
    current_date = datetime.now().date()
    return current_date.strftime("%Y-%m-%d")

def get_current_time():
    """
    Returns the current time in the format HH:MM:SS.
    """
    current_time = datetime.now().time()
    return current_time.strftime("%H:%M:%S")
    

dload_path = '/home/sourabh/rpi_stock_indicator/tmp/dload'
dload_sysfile_name = '/home/sourabh/rpi_stock_indicator/tmp/dload/system_info.xlsx'
dload_stkfile_name = '/home/sourabh/rpi_stock_indicator/tmp/dload/stock_info.xlsx'
tmp_stkfile_name = '/home/sourabh/rpi_stock_indicator/tmp/stock_info.xlsx'

def file_is_under_edit(file_path):
    data = fetch_data_from_excel(file_path, sheet_name="System Info")
    row, col = SystemFields.SYS_EXCEL_STATE.value 
    print(f"System State: {data[row][col]}")
    if data[row][col] == "Running":
        return False
    else:
        return True

def update_all_system_info(file_path, cpu_temp, num_sell_stk, num_buy_stk, num_crt_stk):

    print("Updating all necessary system info")
    data = load_workbook(file_path)
    
    # Sell Status
    if num_sell_stk > 0:
        color = "00FF00"
    else:
        color = "808080"
    r, c = SystemFields.SYS_SELL_STK_FLAG.value
    change_cell_color(data, sheet_name="System Info", row = r, column = c, color=color)
    r, c = SystemFields.SYS_SELL_STK_VAL.value
    update_excel_data(data, sheet_name="System Info", row = r, column = c, new_value=num_sell_stk)
    
    # Buy Status
    if num_buy_stk > 0:
        color = "FFFF00"
    else:
        color = "808080"
    r, c = SystemFields.SYS_BUY_STK_FLAG.value
    change_cell_color(data, sheet_name="System Info", row = r, column = c, color=color)
    r, c = SystemFields.SYS_BUY_STK_VAL.value
    update_excel_data(data, sheet_name="System Info", row = r, column = c, new_value=num_buy_stk)
    data.save(file_path)

    # Crt Status
    if num_crt_stk > 0:
        color = "FF0000"
    else:
        color = "808080"
    r, c = SystemFields.SYS_CRT_STK_FLAG.value
    change_cell_color(data, sheet_name="System Info", row = r, column = c, color=color)
    r, c = SystemFields.SYS_CRT_STK_VAL.value
    update_excel_data(data, sheet_name="System Info", row = r, column = c, new_value=num_crt_stk)

    # CPU Temp
    r, c = SystemFields.SYS_CPU_TEMP.value
    update_excel_data(data, sheet_name="System Info", row = r, column = c, new_value=cpu_temp)

    # CURRENT DATE
    r, c = SystemFields.SYS_CURR_DATE.value
    update_excel_data(data, sheet_name="System Info", row = r, column = c, new_value=get_current_date())

    # CURRENT TIME
    r, c = SystemFields.SYS_CURR_TIME.value
    update_excel_data(data, sheet_name="System Info", row = r, column = c, new_value=get_current_time())

    data.save(file_path)    
    

def process_wishlist(workbook, buy):
    data = fetch_data_from_sheet(workbook, "Wishlist")
    idx_col = StkWishList.S_NO.value
    tkr_col = StkWishList.TCKR.value
    cur_col = StkWishList.CURRENT.value
    type_col = StkWishList.TYPE.value
    trgt_col = StkWishList.TARGET.value
    
    for i in range(MAX_ALLOWED_ROWS):
      if i == 0:
        continue
        
      idx = data[i][idx_col]
      tkr = data[i][tkr_col]
      tkr_type = data[i][type_col]
      target = data[i][trgt_col]
      
      if tkr is None:
        break;
        
      stock_price = get_stock_price(tkr, tkr_type)
      print(f"idx {idx} tkr {tkr} price {stock_price} target {target}")
      if "Error" in str(stock_price):
        update_excel_data(workbook, "Wishlist", i, cur_col, "Invalid TICKR")
      else:
        update_excel_data(workbook, "Wishlist", i, cur_col, stock_price)
        
      color="FFFFFF"
      if target is None:
        print("Target is not set...")
      else:
        print("******* TARGET REACHED")
        if stock_price < target:
          color="FFFF00"
          buy[0] += 1
          
      change_row_color(workbook, "Wishlist", i, color)
          
    print(f"buy = {buy}")
          
def process_portfolio(workbook, sell, crt):
    data = fetch_data_from_sheet(workbook, "Portfolio")
    idx_col = StkPortfolio.S_NO.value
    tkr_col = StkPortfolio.TCKR.value
    cur_col = StkPortfolio.CURRENT.value
    type_col = StkPortfolio.TYPE.value
    trgt_col = StkPortfolio.TARGET.value
    
    for i in range(MAX_ALLOWED_ROWS):
      if i == 0:
        continue
        
      idx = data[i][idx_col]
      tkr = data[i][tkr_col]
      tkr_type = data[i][type_col]
      target = data[i][trgt_col]
      
      if tkr is None:
        break;
        
      stock_price = get_stock_price(tkr, tkr_type)
      print(f"idx {idx} tkr {tkr} price {stock_price} target {target}")
      if "Error" in str(stock_price):
        update_excel_data(workbook, "Portfolio", i, cur_col, "Invalid TICKR")
      else:
        update_excel_data(workbook, "Portfolio", i, cur_col, stock_price)
        
      color="FFFFFF"
      if target is None:
        print("Target is not set...")
      else:
        print("******* TARGET REACHED")
        if stock_price > target:
          color="00FF00"
          sell[0] += 1
          
      change_row_color(workbook, "Portfolio", i, color)
          
    print(f"sell = {sell}")        
    

def monitor_stock_market(inst):
    
    print("Monitor stock market...")
    download_file_from_gdrive(dload_path)
    if file_is_under_edit(dload_sysfile_name) is True:
        print("File under edit")
        inst.to_system_queue.put((TASK_SYSTEM_DEFAULT, "Excel under edit"))
        return
        

    copy_file(dload_stkfile_name, tmp_stkfile_name)
    
    data = load_workbook(tmp_stkfile_name)

    buy = [0]
    process_wishlist(data, buy)    
    # ***** Raise the Buy flag if required *****
    print(f"--------- {inst.num_buy_stk} {buy}")
    if inst.num_buy_stk == 0 and buy[0] > 0:
      inst.to_system_queue.put((TASK_SYSTEM_STK_BUY, "Buy flag set..."))
    elif inst.num_buy_stk > 0 and buy[0] == 0:
      inst.to_system_queue.put((TASK_SYSTEM_STK_BUY_CLR, "Buy flag cleared..."))      
    inst.num_buy_stk = buy[0]
    
    
    sell = [0]
    crt = [0]
    process_portfolio(data, sell, crt)    
    # ***** Raise the sell flag if required *****
    print(f"--------- {inst.num_sell_stk} {sell}")
    if inst.num_sell_stk == 0 and sell[0] > 0:
      inst.to_system_queue.put((TASK_SYSTEM_STK_SELL, "Sell flag set..."))
    elif inst.num_sell_stk > 0 and sell[0] == 0:
      inst.to_system_queue.put((TASK_SYSTEM_STK_SELL_CLR, "Sell flag cleared..."))      
    inst.num_sell_stk = sell[0]
    
    data.save(tmp_stkfile_name)
    
    download_file_from_gdrive(dload_path)
    if file_is_under_edit(dload_sysfile_name) is True:
        print("File under edit")
        inst.to_system_queue.put((TASK_SYSTEM_DEFAULT, "Excel under edit"))
        return

    update_all_system_info(dload_sysfile_name, inst.cpu_temp, inst.num_sell_stk, inst.num_buy_stk, inst.num_crt_stk)
    
    upload_file_to_gdrive(dload_sysfile_name)
    upload_file_to_gdrive(tmp_stkfile_name)
    inst.to_system_queue.put((TASK_SYSTEM_RUNNING, "stocks monitor works fine"))



